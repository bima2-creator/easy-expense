import os
import io
import csv
import re
import time
import uuid
import json
import base64
import zipfile
import logging
from pathlib import Path
from datetime import datetime, timezone
from typing import List, Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Query, Depends, Header
from fastapi.responses import Response, StreamingResponse
from starlette.concurrency import run_in_threadpool
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from dotenv import load_dotenv

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as rl_colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

APP_NAME = "easy-expense"

# ---------------------------------------------------------------------------
# Simple shared-secret auth (single-user app, no login flow needed).
# Backend ini publik lewat Cloudflare Tunnel tanpa proteksi apa pun sebelumnya —
# setiap request ke /api/* sekarang wajib bawa header X-API-Key yang cocok.
# ---------------------------------------------------------------------------
API_KEY = os.environ.get('API_KEY')
if not API_KEY:
    raise RuntimeError("API_KEY belum diset di backend/.env — wajib diisi supaya API tidak terbuka untuk umum")


async def require_api_key(x_api_key: Optional[str] = Header(default=None)):
    if x_api_key != API_KEY:
        raise HTTPException(status_code=401, detail="API key tidak valid")


api_router = APIRouter(prefix="/api", dependencies=[Depends(require_api_key)])

# Local disk storage for receipt images (replaces Emergent Object Storage)
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = FastAPI()

DEFAULT_CATEGORIES = [
    {"name": "Makan", "icon": "restaurant-outline", "color": "#C85A44"},
    {"name": "Minum", "icon": "cafe-outline", "color": "#4A7C7C"},
    {"name": "Entertain", "icon": "film-outline", "color": "#B5573F"},
    {"name": "Fuel", "icon": "car-sport-outline", "color": "#8F6B5B"},
    {"name": "Toll", "icon": "trail-sign-outline", "color": "#5B8F6B"},
    {"name": "Office Supplies", "icon": "briefcase-outline", "color": "#5B6C8F"},
]
CUSTOM_COLORS = ["#4A7C59", "#D9943B", "#7A5B8F", "#5B6C8F", "#8F7A5B", "#4A7C7C", "#B5573F"]


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def rupiah(amount) -> str:
    try:
        return "Rp " + f"{int(round(float(amount or 0))):,}".replace(",", ".")
    except Exception:
        return "Rp 0"


# ---------------------------------------------------------------------------
# Object Storage (local disk)
# ---------------------------------------------------------------------------
def _safe_local_path(path: str) -> Path:
    # path looks like "easy-expense/uploads/default/<uuid>.jpg"
    full = (UPLOAD_DIR / path).resolve()
    if not str(full).startswith(str(UPLOAD_DIR.resolve())):
        raise ValueError("Invalid path")
    return full


def put_object(path: str, data: bytes, content_type: str) -> dict:
    full = _safe_local_path(path)
    full.parent.mkdir(parents=True, exist_ok=True)
    full.write_bytes(data)
    return {"path": path}


def get_object(path: str):
    full = _safe_local_path(path)
    if not full.exists():
        raise FileNotFoundError(path)
    ext = full.suffix.lower().lstrip(".")
    ctype = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp"}.get(ext, "application/octet-stream")
    return full.read_bytes(), ctype


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class Category(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    icon: str = "pricetag-outline"
    color: str = "#4A7C59"
    is_default: bool = False
    order: int = 0
    created_at: str = Field(default_factory=now_iso)


class CategoryCreate(BaseModel):
    name: str
    icon: Optional[str] = None


class CategoryUpdate(BaseModel):
    name: str
    icon: Optional[str] = None


class CategoryReorder(BaseModel):
    ids: List[str]  # urutan id kategori sesuai keinginan user, dari atas ke bawah


class Project(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    client: Optional[str] = ""
    color: Optional[str] = "#4A7C59"
    parent_id: Optional[str] = None  # kalau diisi, project ini adalah sub-proyek dari project lain
    is_paid: bool = False  # untuk pengeluaran yang langsung di bawah proyek/sub-proyek (tanpa WO)
    paid_at: Optional[str] = None
    created_at: str = Field(default_factory=now_iso)


class ProjectCreate(BaseModel):
    name: str
    client: Optional[str] = ""
    color: Optional[str] = "#4A7C59"
    parent_id: Optional[str] = None


class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    client: Optional[str] = None
    color: Optional[str] = None
    is_paid: Optional[bool] = None


class WorkOrder(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    name: str
    is_paid: bool = False
    paid_at: Optional[str] = None
    created_at: str = Field(default_factory=now_iso)


class WorkOrderCreate(BaseModel):
    project_id: str
    name: str


class WorkOrderUpdate(BaseModel):
    name: Optional[str] = None
    is_paid: Optional[bool] = None


class Expense(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vendor: str = "Vendor Tidak Diketahui"
    amount: float = 0.0
    date: str = Field(default_factory=lambda: datetime.now(timezone.utc).date().isoformat())
    category: str = "Makan"
    project_id: Optional[str] = None
    work_order_id: Optional[str] = None
    notes: Optional[str] = ""
    receipt_path: Optional[str] = None
    is_billable: bool = False
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)


class ExpenseCreate(BaseModel):
    vendor: str = "Vendor Tidak Diketahui"
    amount: float = 0.0
    date: Optional[str] = None
    category: str = "Makan"
    project_id: Optional[str] = None
    work_order_id: Optional[str] = None
    notes: Optional[str] = ""
    receipt_path: Optional[str] = None
    is_billable: bool = False


class ExpenseUpdate(BaseModel):
    vendor: Optional[str] = None
    amount: Optional[float] = None
    date: Optional[str] = None
    category: Optional[str] = None
    project_id: Optional[str] = None
    work_order_id: Optional[str] = None
    notes: Optional[str] = None
    is_billable: Optional[bool] = None


# ---------------------------------------------------------------------------
# Receipt extraction: OCR.space (free tier, no local ML needed) + heuristic parsing
# ---------------------------------------------------------------------------
OCR_SPACE_API_KEY = os.environ.get('OCR_SPACE_API_KEY')
OCR_SPACE_URL = "https://api.ocr.space/parse/image"

# Kata kunci per kategori untuk menebak kategori dari isi struk (Bahasa Indonesia + umum)
CATEGORY_KEYWORDS = {
    "Makan": ["restoran", "resto", "warung", "rumah makan", "cafe", "kafe", "nasi", "ayam",
              "bakso", "mie", "soto", "padang", "kfc", "mcdonald", "mcd", "pizza", "burger"],
    "Minum": ["kopi", "coffee", "starbucks", "juice", "jus", "boba", "teh", "minuman", "cafe"],
    "Entertain": ["cinema", "bioskop", "xxi", "cgv", "karaoke", "game", "billiard", "tiket"],
    "Fuel": ["pertamina", "shell", "bensin", "pertalite", "pertamax", "solar", "spbu", "bbm"],
    "Toll": ["tol", "toll", "e-toll", "gerbang"],
    "Office Supplies": ["atk", "stationery", "kertas", "printer", "tinta", "gramedia", "toko buku"],
}

# Kata yang menandakan baris "grand total" (prioritas tinggi) vs subtotal (prioritas rendah)
TOTAL_HINTS = ["grand total", "total bayar", "total belanja", "total tagihan", "jumlah bayar", "nominal", "total"]
SKIP_TOTAL_HINTS = ["subtotal", "sub total", "kembali", "kembalian", "tunai", "cash", "diskon", "discount", "dp ",
                     "rrn", "approval", "terminal", "reff", "ref no", "no ref", "invoice", "trace", "kode",
                     "cn:", "saldo", "e-money balance", "sisa saldo", "balance"]
MAX_REASONABLE_AMOUNT = 999_999_999  # ceiling wajar untuk satu transaksi (di bawah 1 miliar Rupiah)

MONTHS_ID = {
    "jan": 1, "januari": 1, "feb": 2, "februari": 2, "mar": 3, "maret": 3, "apr": 4, "april": 4,
    "mei": 5, "may": 5, "jun": 6, "juni": 6, "jul": 7, "juli": 7, "agu": 8, "agt": 8, "agustus": 8, "aug": 8,
    "sep": 9, "september": 9, "okt": 10, "oktober": 10, "oct": 10, "nov": 11, "november": 11,
    "des": 12, "desember": 12, "dec": 12,
}


def ocr_space_read_text(image_bytes: bytes, content_type: str) -> str:
    if not OCR_SPACE_API_KEY:
        raise RuntimeError("OCR_SPACE_API_KEY belum diset di backend/.env")
    ext = "jpg"
    if "png" in content_type:
        ext = "png"

    last_error = None
    for attempt in range(3):  # coba sampai 3x, karena OCR.space free tier kadang 503 sesaat
        try:
            resp = requests.post(
                OCR_SPACE_URL,
                files={"file": (f"receipt.{ext}", image_bytes, content_type)},
                data={"apikey": OCR_SPACE_API_KEY, "language": "eng", "OCREngine": "2", "scale": "true"},
                timeout=60,
            )
            if resp.status_code in (503, 502, 504) and attempt < 2:
                last_error = f"{resp.status_code} Server Error"
                time.sleep(2)
                continue
            resp.raise_for_status()
            result = resp.json()
            if result.get("IsErroredOnProcessing"):
                raise RuntimeError(str(result.get("ErrorMessage") or "OCR gagal memproses gambar"))
            parsed = result.get("ParsedResults") or []
            return parsed[0].get("ParsedText", "") if parsed else ""
        except requests.exceptions.RequestException as e:
            last_error = str(e)
            if attempt < 2:
                time.sleep(2)
                continue
            raise
    raise RuntimeError(last_error or "OCR gagal setelah beberapa percobaan")


def _parse_amount_token(tok: str) -> Optional[float]:
    digits_only = re.sub(r"[^\d]", "", tok)
    # Nomor referensi/RRN/terminal ID biasanya sangat panjang (10+ digit) -> bukan nominal uang
    if len(digits_only) > 9:
        return None
    raw = re.sub(r"[^\d.,]", "", tok)
    if not raw:
        return None
    # Buang bagian sen di akhir kalau ada, mis. "798.770,00" atau "798.770.00" -> "798.770"
    raw = re.sub(r"[.,]\d{2}$", "", raw)
    # Sisa titik/koma dianggap pemisah ribuan, hapus semua
    val_str = re.sub(r"[.,]", "", raw)
    if not val_str:
        return None
    try:
        val = float(val_str)
    except ValueError:
        return None
    if val > MAX_REASONABLE_AMOUNT:
        return None
    return val


RP_PREFIX_RE = re.compile(r"rp\.?\s*([\d]+(?:[.,]\d+)*)", re.IGNORECASE)
BALANCE_LINE_HINTS = ["cn:", "saldo", "e-money balance", "sisa saldo", "balance"]
TOLL_LINE_HINTS = ["e-toll", "e toll", "gol-", "gol "]


def guess_amount(lines: List[str]) -> float:
    best = None
    for line in lines:
        low = line.lower()
        if any(h in low for h in SKIP_TOTAL_HINTS):
            continue
        if any(h in low for h in TOTAL_HINTS):
            nums = re.findall(r"\d[\d.,]*\d|\d", line)
            for n in nums:
                val = _parse_amount_token(n)
                if val and val > 0:
                    best = val  # ambil angka terakhir di baris (biasanya nominalnya)
    if best:
        return best

    # Prioritas khusus: struk e-Toll, format umum "Gol-1 e-Toll MANDIRI 21500"
    # (nomor di baris ini adalah biaya tol, BUKAN saldo kartu yang biasa muncul di baris terpisah)
    for line in lines:
        low = line.lower()
        if any(h in low for h in TOLL_LINE_HINTS):
            nums = re.findall(r"\d[\d.,]*\d|\d", line)
            for n in reversed(nums):  # ambil angka terakhir di baris itu
                val = _parse_amount_token(n)
                if val and val >= 100:
                    return val

    # Prioritas kedua: angka yang diawali "Rp" (struk transfer/QRIS jarang punya kata "Total",
    # tapi nominal utama biasanya satu-satunya angka berprefix "Rp" tanpa embel-embel referensi)
    full_text = "\n".join(lines)
    rp_matches = []
    for line in lines:
        low = line.lower()
        if any(h in low for h in SKIP_TOTAL_HINTS):
            continue
        for m in RP_PREFIX_RE.finditer(line):
            val = _parse_amount_token(m.group(1))
            if val and val >= 100:
                rp_matches.append(val)
    if rp_matches:
        return rp_matches[0]  # ambil kemunculan pertama (biasanya nominal utama ditampilkan di awal)

    # fallback terakhir: angka terbesar di seluruh struk (kemungkinan grand total)
    all_vals = []
    for line in lines:
        low = line.lower()
        if any(h in low for h in SKIP_TOTAL_HINTS):
            continue
        for n in re.findall(r"\d[\d.,]*\d|\d", line):
            val = _parse_amount_token(n)
            if val and val >= 100:
                all_vals.append(val)
    return max(all_vals) if all_vals else 0.0


EXPIRY_KEYWORDS = ["ed:", "ed ", "exp:", "exp ", "expired", "kadaluarsa", "kadaluwarsa", "best before"]
DATE_KEYWORDS = ["tanggal", "tgl", "date"]


def _try_date_patterns(text: str) -> Optional[str]:
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, mo, d).date().isoformat()
        except ValueError:
            pass
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", text)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return datetime(y, mo, d).date().isoformat()
        except ValueError:
            pass
    m = re.search(r"(\d{1,2})\s+([A-Za-z]{3,9})\s+(\d{4})", text, re.IGNORECASE)
    if m:
        d = int(m.group(1))
        mo = MONTHS_ID.get(m.group(2).lower())
        y = int(m.group(3))
        if mo:
            try:
                return datetime(y, mo, d).date().isoformat()
            except ValueError:
                pass
    return None


def guess_date(text: str) -> str:
    today = datetime.now(timezone.utc).date()
    all_lines = text.splitlines()
    # Buang baris yang berisi tanggal kadaluarsa produk (bukan tanggal transaksi)
    clean_lines = [l for l in all_lines if not any(kw in l.lower() for kw in EXPIRY_KEYWORDS)]

    # Prioritas 1: baris yang eksplisit mengandung kata "tanggal/tgl/date"
    for line in clean_lines:
        if any(kw in line.lower() for kw in DATE_KEYWORDS):
            found = _try_date_patterns(line)
            if found:
                return found

    # Prioritas 2: cari di seluruh teks yang sudah dibersihkan dari baris kadaluarsa
    found = _try_date_patterns("\n".join(clean_lines))
    if found:
        found_date = datetime.fromisoformat(found).date()
        # Tolak tanggal yang jauh di masa depan (>30 hari) - kemungkinan salah tangkap
        if (found_date - today).days <= 30:
            return found

    return today.isoformat()


def guess_vendor(lines: List[str]) -> str:
    for line in lines[:6]:
        clean = line.strip()
        letters = sum(ch.isalpha() for ch in clean)
        digits = sum(ch.isdigit() for ch in clean)
        if letters >= 3 and letters > digits and len(clean) <= 40:
            return clean.title() if clean.isupper() else clean
    return "Vendor Tidak Diketahui"


def guess_category(text: str, category_names: List[str]) -> str:
    low = text.lower()
    for cat_name in category_names:
        for kw in CATEGORY_KEYWORDS.get(cat_name, []):
            if kw in low:
                return cat_name
    return category_names[0] if category_names else "Makan"


async def extract_receipt(image_bytes: bytes, content_type: str, category_names: List[str]) -> dict:
    raw_text = await run_in_threadpool(ocr_space_read_text, image_bytes, content_type)
    lines = [l.strip() for l in raw_text.splitlines() if l.strip()]

    return {
        "vendor": guess_vendor(lines),
        "amount": guess_amount(lines),
        "date": guess_date(raw_text),
        "category": guess_category(raw_text, category_names),
        "notes": "",
    }


# ---------------------------------------------------------------------------
# Categories
# ---------------------------------------------------------------------------
async def ensure_default_categories():
    count = await db.categories.count_documents({})
    if count == 0:
        for i, c in enumerate(DEFAULT_CATEGORIES):
            cat = Category(name=c["name"], icon=c["icon"], color=c["color"], is_default=True, order=i)
            await db.categories.insert_one(cat.dict())


@api_router.get("/categories", response_model=List[Category])
async def list_categories():
    docs = await db.categories.find().sort("order", 1).to_list(1000)
    return [Category(**d) for d in docs]


@api_router.post("/categories", response_model=Category)
async def create_category(body: CategoryCreate):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nama kategori wajib diisi")
    existing = await db.categories.find_one({"name": name})
    if existing:
        raise HTTPException(status_code=400, detail="Kategori sudah ada")
    n = await db.categories.count_documents({})
    cat = Category(name=name, icon=body.icon or "pricetag-outline", color=CUSTOM_COLORS[n % len(CUSTOM_COLORS)], order=n)
    await db.categories.insert_one(cat.dict())
    return cat


@api_router.put("/categories/reorder")
async def reorder_categories(body: CategoryReorder):
    for i, cid in enumerate(body.ids):
        await db.categories.update_one({"id": cid}, {"$set": {"order": i}})
    return {"ok": True}


@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(category_id: str, body: CategoryUpdate):
    doc = await db.categories.find_one({"id": category_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Kategori tidak ditemukan")
    new_name = body.name.strip()
    old_name = doc["name"]
    update = {"name": new_name}
    if body.icon:
        update["icon"] = body.icon
    await db.categories.update_one({"id": category_id}, {"$set": update})
    # propagate rename to expenses
    if new_name != old_name:
        await db.expenses.update_many({"category": old_name}, {"$set": {"category": new_name}})
    doc = await db.categories.find_one({"id": category_id})
    return Category(**doc)


@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str):
    doc = await db.categories.find_one({"id": category_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Kategori tidak ditemukan")
    await db.categories.delete_one({"id": category_id})
    return {"ok": True}


# ---------------------------------------------------------------------------
# Projects
# ---------------------------------------------------------------------------
async def _project_rollup(project_id: str) -> dict:
    """Hitung total & jumlah transaksi milik proyek ini SAMPAI KE SEMUA SUB-PROYEK
    di bawahnya (rekursif). Pengeluaran yang menempel ke work order tetap ikut
    terhitung di sini karena field project_id-nya tetap menunjuk ke proyek/sub-
    proyek induk WO tersebut, bukan cuma pengeluaran langsung tanpa WO."""
    agg = await db.expenses.aggregate([
        {"$match": {"project_id": project_id}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}, "count": {"$sum": 1}}},
    ]).to_list(1)
    total = agg[0]["total"] if agg else 0
    count = agg[0]["count"] if agg else 0
    children = await db.projects.find({"parent_id": project_id}).to_list(1000)
    for c in children:
        child = await _project_rollup(c["id"])
        total += child["total"]
        count += child["count"]
    return {"total": round(total, 2), "count": count}


@api_router.get("/projects")
async def list_projects(parent_id: Optional[str] = Query(None), all: bool = Query(False)):
    # Default (tanpa parent_id): cuma project level teratas (parent_id kosong) —
    # ini termasuk semua project lama sebelum fitur sub-proyek ada, karena query
    # {"parent_id": None} di MongoDB juga cocok dengan dokumen yang field-nya
    # belum ada sama sekali (tidak perlu migrasi data).
    # Kirim ?parent_id=<id project> untuk ambil daftar sub-proyek dari project itu.
    # Kirim ?all=true untuk ambil semua project di semua level sekaligus.
    if all:
        q = {}
    elif parent_id:
        q = {"parent_id": parent_id}
    else:
        q = {"parent_id": None}
    docs = await db.projects.find(q).sort("created_at", -1).to_list(1000)
    result = []
    for d in docs:
        p = Project(**d).dict()
        wo_count = await db.work_orders.count_documents({"project_id": p["id"]})
        sub_count = await db.projects.count_documents({"parent_id": p["id"]})
        rollup = await _project_rollup(p["id"])
        p["work_order_count"] = wo_count
        p["sub_project_count"] = sub_count
        p["total"] = rollup["total"]
        p["expense_count"] = rollup["count"]
        result.append(p)
    return result


@api_router.get("/projects/{project_id}")
async def get_project(project_id: str):
    doc = await db.projects.find_one({"id": project_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")
    p = Project(**doc).dict()
    wo_count = await db.work_orders.count_documents({"project_id": project_id})
    sub_count = await db.projects.count_documents({"parent_id": project_id})
    rollup = await _project_rollup(project_id)
    p["work_order_count"] = wo_count
    p["sub_project_count"] = sub_count
    p["total"] = rollup["total"]
    p["expense_count"] = rollup["count"]
    return p


@api_router.post("/projects", response_model=Project)
async def create_project(body: ProjectCreate):
    proj = Project(**body.dict())
    await db.projects.insert_one(proj.dict())
    return proj


@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, body: ProjectUpdate):
    doc = await db.projects.find_one({"id": project_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")
    updates = {k: v for k, v in body.dict(exclude_unset=True).items() if k != "is_paid"}
    if body.is_paid is not None:
        updates["is_paid"] = body.is_paid
        updates["paid_at"] = now_iso() if body.is_paid else None
    if updates:
        await db.projects.update_one({"id": project_id}, {"$set": updates})
    doc = await db.projects.find_one({"id": project_id})
    return Project(**doc)


@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str):
    res = await db.projects.delete_one({"id": project_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")
    # detach work orders and expenses
    await db.work_orders.delete_many({"project_id": project_id})
    await db.expenses.update_many({"project_id": project_id}, {"$set": {"project_id": None, "work_order_id": None}})
    # sub-proyek TIDAK ikut terhapus — naik jadi project level teratas supaya
    # datanya (work order, pengeluaran, dst di dalamnya) tetap aman
    await db.projects.update_many({"parent_id": project_id}, {"$set": {"parent_id": None}})
    return {"ok": True}


# ---------------------------------------------------------------------------
# Work Orders
# ---------------------------------------------------------------------------
@api_router.get("/work-orders")
async def list_work_orders(project_id: Optional[str] = Query(None), is_paid: Optional[bool] = Query(None)):
    q = {}
    if project_id:
        q["project_id"] = project_id
    if is_paid is not None:
        q["is_paid"] = is_paid
    docs = await db.work_orders.find(q).sort("created_at", -1).to_list(1000)
    result = []
    for d in docs:
        wo = WorkOrder(**d).dict()
        agg = await db.expenses.aggregate([
            {"$match": {"work_order_id": wo["id"]}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}, "count": {"$sum": 1}}},
        ]).to_list(1)
        wo["total"] = round(agg[0]["total"], 2) if agg else 0
        wo["expense_count"] = agg[0]["count"] if agg else 0
        result.append(wo)
    return result


@api_router.get("/work-orders/{wo_id}")
async def get_work_order(wo_id: str):
    doc = await db.work_orders.find_one({"id": wo_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Work order tidak ditemukan")
    return WorkOrder(**doc)


@api_router.post("/work-orders", response_model=WorkOrder)
async def create_work_order(body: WorkOrderCreate):
    proj = await db.projects.find_one({"id": body.project_id})
    if not proj:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")
    wo = WorkOrder(**body.dict())
    await db.work_orders.insert_one(wo.dict())
    return wo


@api_router.put("/work-orders/{wo_id}", response_model=WorkOrder)
async def update_work_order(wo_id: str, body: WorkOrderUpdate):
    doc = await db.work_orders.find_one({"id": wo_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Work order tidak ditemukan")
    update_fields = {}
    if body.name is not None:
        update_fields["name"] = body.name
    if body.is_paid is not None:
        update_fields["is_paid"] = body.is_paid
        update_fields["paid_at"] = now_iso() if body.is_paid else None
    if update_fields:
        await db.work_orders.update_one({"id": wo_id}, {"$set": update_fields})
    doc = await db.work_orders.find_one({"id": wo_id})
    return WorkOrder(**doc)


@api_router.delete("/work-orders/{wo_id}")
async def delete_work_order(wo_id: str):
    res = await db.work_orders.delete_one({"id": wo_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Work order tidak ditemukan")
    await db.expenses.update_many({"work_order_id": wo_id}, {"$set": {"work_order_id": None}})
    return {"ok": True}


# ---------------------------------------------------------------------------
# Scan & Files
# ---------------------------------------------------------------------------
@api_router.post("/scan")
async def scan_receipt(file: UploadFile = File(...)):
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="File kosong")

    ext = "jpg"
    ctype = file.content_type or "image/jpeg"
    if "png" in ctype:
        ext = "png"
    elif "webp" in ctype:
        ext = "webp"

    path = f"{APP_NAME}/uploads/default/{uuid.uuid4()}.{ext}"
    try:
        await run_in_threadpool(put_object, path, raw, ctype)
    except Exception as e:
        logger.error(f"Storage upload failed: {e}")
        raise HTTPException(status_code=502, detail="Gagal menyimpan struk")

    cat_docs = await db.categories.find().sort("created_at", 1).to_list(1000)
    category_names = [c["name"] for c in cat_docs] or [c["name"] for c in DEFAULT_CATEGORIES]

    try:
        extracted = await extract_receipt(raw, ctype, category_names)
    except Exception as e:
        logger.error(f"OCR extraction failed: {e}")
        extracted = {
            "vendor": "Vendor Tidak Diketahui", "amount": 0,
            "date": datetime.now(timezone.utc).date().isoformat(),
            "category": category_names[0], "notes": "",
        }
        return {"receipt_path": path, "extracted": extracted, "extraction_failed": True, "duplicate": None}

    # Deteksi kemungkinan struk duplikat. Dicocokkan berdasarkan tanggal + jumlah
    # (toleransi kecil) — BUKAN vendor persis, karena hasil OCR untuk teks vendor
    # sering meleset dikit antar-scan (mis. struk transfer bank: "m-Transfer" vs
    # "n-Transfer" salah baca huruf), jadi vendor persis terlalu ketat dan malah
    # sering gagal mendeteksi struk yang sebenarnya sama. Ini cuma peringatan,
    # bukan blokir keras — user tetap boleh lanjut kalau memang bukan duplikat.
    duplicate = None
    amount = extracted.get("amount") or 0
    date_val = extracted.get("date")
    if amount and date_val:
        dup = await db.expenses.find_one({
            "date": date_val,
            "amount": {"$gte": amount - 1, "$lte": amount + 1},
        })
        if dup:
            duplicate = {
                "expense_id": dup["id"], "vendor": dup.get("vendor"),
                "date": dup.get("date"), "amount": dup.get("amount"),
                "category": dup.get("category"),
            }

    return {"receipt_path": path, "extracted": extracted, "extraction_failed": False, "duplicate": duplicate}


@api_router.get("/files/{path:path}")
async def get_file(path: str):
    try:
        content, ctype = await run_in_threadpool(get_object, path)
    except Exception as e:
        logger.error(f"Storage read failed: {e}")
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    return Response(content=content, media_type=ctype, headers={"Cache-Control": "public, max-age=86400"})


# ---------------------------------------------------------------------------
# Expenses
# ---------------------------------------------------------------------------
async def _scope_is_paid(project_id: Optional[str], work_order_id: Optional[str]) -> bool:
    """Cek apakah WO atau Proyek/Sub-Proyek tujuan sudah ditandai lunas.
    WO diprioritaskan kalau ada (expense di bawah WO ikut status WO-nya,
    bukan status proyek induknya)."""
    if work_order_id:
        wo = await db.work_orders.find_one({"id": work_order_id})
        return bool(wo and wo.get("is_paid"))
    if project_id:
        proj = await db.projects.find_one({"id": project_id})
        return bool(proj and proj.get("is_paid"))
    return False


@api_router.post("/expenses", response_model=Expense)
async def create_expense(body: ExpenseCreate):
    if await _scope_is_paid(body.project_id, body.work_order_id):
        raise HTTPException(status_code=400, detail="WO/Proyek ini sudah lunas — tidak bisa menambah pengeluaran baru di sini.")
    data = body.dict()
    if not data.get("date"):
        data.pop("date", None)
    exp = Expense(**{k: v for k, v in data.items() if v is not None or k in ("project_id", "work_order_id")})
    await db.expenses.insert_one(exp.dict())
    return exp


@api_router.get("/expenses")
async def list_expenses(
    category: Optional[str] = Query(None),
    project_id: Optional[str] = Query(None),
    work_order_id: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
):
    q = {}
    if category and category != "Semua":
        q["category"] = category
    if project_id:
        q["project_id"] = project_id
    if work_order_id:
        q["work_order_id"] = work_order_id
    if search:
        q["vendor"] = {"$regex": search, "$options": "i"}
    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to
        q["date"] = date_q
    docs = await db.expenses.find(q).sort("date", -1).to_list(2000)

    wo_ids = {d["work_order_id"] for d in docs if d.get("work_order_id")}
    proj_ids = {d["project_id"] for d in docs if d.get("project_id") and not d.get("work_order_id")}
    wo_paid = {}
    if wo_ids:
        for w in await db.work_orders.find({"id": {"$in": list(wo_ids)}}).to_list(2000):
            wo_paid[w["id"]] = bool(w.get("is_paid"))
    proj_paid = {}
    if proj_ids:
        for p in await db.projects.find({"id": {"$in": list(proj_ids)}}).to_list(2000):
            proj_paid[p["id"]] = bool(p.get("is_paid"))

    result = []
    for d in docs:
        e = Expense(**d).dict()
        if d.get("work_order_id"):
            e["locked"] = wo_paid.get(d["work_order_id"], False)
        elif d.get("project_id"):
            e["locked"] = proj_paid.get(d["project_id"], False)
        else:
            e["locked"] = False
        result.append(e)
    return result


@api_router.get("/expenses/{expense_id}")
async def get_expense(expense_id: str):
    doc = await db.expenses.find_one({"id": expense_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Pengeluaran tidak ditemukan")
    e = Expense(**doc).dict()
    e["locked"] = await _scope_is_paid(doc.get("project_id"), doc.get("work_order_id"))
    return e


@api_router.put("/expenses/{expense_id}", response_model=Expense)
async def update_expense(expense_id: str, body: ExpenseUpdate):
    doc = await db.expenses.find_one({"id": expense_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Pengeluaran tidak ditemukan")
    if await _scope_is_paid(doc.get("project_id"), doc.get("work_order_id")):
        raise HTTPException(status_code=403, detail="Pengeluaran ini terkunci karena WO/Proyeknya sudah lunas.")
    updates = {k: v for k, v in body.dict(exclude_unset=True).items()}
    target_project = updates.get("project_id", doc.get("project_id"))
    target_wo = updates.get("work_order_id", doc.get("work_order_id"))
    if (target_project, target_wo) != (doc.get("project_id"), doc.get("work_order_id")):
        if await _scope_is_paid(target_project, target_wo):
            raise HTTPException(status_code=400, detail="Tidak bisa memindahkan pengeluaran ke WO/Proyek yang sudah lunas.")
    updates["updated_at"] = now_iso()
    await db.expenses.update_one({"id": expense_id}, {"$set": updates})
    doc = await db.expenses.find_one({"id": expense_id})
    return Expense(**doc)


@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str):
    doc = await db.expenses.find_one({"id": expense_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Pengeluaran tidak ditemukan")
    if await _scope_is_paid(doc.get("project_id"), doc.get("work_order_id")):
        raise HTTPException(status_code=403, detail="Pengeluaran ini terkunci karena WO/Proyeknya sudah lunas.")
    await db.expenses.delete_one({"id": expense_id})
    return {"ok": True}


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------
@api_router.get("/reports/summary")
async def reports_summary(period: str = Query("month"), date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None)):
    docs = await db.expenses.find().to_list(5000)
    now = datetime.now(timezone.utc)

    def in_period(d):
        try:
            dt = datetime.fromisoformat(d["date"])
        except Exception:
            return True
        if period == "custom":
            if date_from and dt.date() < datetime.fromisoformat(date_from).date():
                return False
            if date_to and dt.date() > datetime.fromisoformat(date_to).date():
                return False
            return True
        if period == "week":
            return (now.date() - dt.date()).days <= 7
        if period == "month":
            return dt.year == now.year and dt.month == now.month
        if period == "quarter":
            return dt.year == now.year and ((dt.month - 1) // 3) == ((now.month - 1) // 3)
        if period == "year":
            return dt.year == now.year
        return True

    filtered = [d for d in docs if in_period(d)]
    total = sum(float(d.get("amount") or 0) for d in filtered)

    by_cat = {}
    for d in filtered:
        c = d.get("category", "Lainnya")
        by_cat[c] = by_cat.get(c, 0) + float(d.get("amount") or 0)
    by_category = sorted(
        [{"category": k, "amount": round(v, 2)} for k, v in by_cat.items()],
        key=lambda x: x["amount"], reverse=True,
    )

    projects = {p["id"]: p["name"] for p in await db.projects.find().to_list(1000)}
    by_proj = {}
    for d in filtered:
        pid = d.get("project_id")
        if pid:
            by_proj[pid] = by_proj.get(pid, 0) + float(d.get("amount") or 0)
    by_project = sorted(
        [{"id": k, "name": projects.get(k, "Proyek"), "amount": round(v, 2)} for k, v in by_proj.items()],
        key=lambda x: x["amount"], reverse=True,
    )

    trend = []
    for i in range(5, -1, -1):
        m = (now.month - i - 1) % 12 + 1
        y = now.year + ((now.month - i - 1) // 12)
        msum = 0.0
        for d in docs:
            try:
                dt = datetime.fromisoformat(d["date"])
                if dt.year == y and dt.month == m:
                    msum += float(d.get("amount") or 0)
            except Exception:
                continue
        trend.append({"label": datetime(y, m, 1).strftime("%b"), "amount": round(msum, 2)})

    return {
        "period": period,
        "total": round(total, 2),
        "count": len(filtered),
        "by_category": by_category,
        "by_project": by_project,
        "trend": trend,
    }


@api_router.get("/reports/export-xlsx")
async def export_xlsx():
    docs = await db.expenses.find().sort("date", -1).to_list(5000)
    projects = {p["id"]: p["name"] for p in await db.projects.find().to_list(1000)}
    project_paid = {p["id"]: bool(p.get("is_paid")) for p in await db.projects.find().to_list(1000)}
    work_orders = {w["id"]: w["name"] for w in await db.work_orders.find().to_list(2000)}
    wo_paid = {w["id"]: bool(w.get("is_paid")) for w in await db.work_orders.find().to_list(2000)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Pengeluaran"

    headers = ["Tanggal", "Vendor", "Kategori", "Proyek", "Work Order", "Jumlah (Rp)", "Billable", "Lunas", "Catatan"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1C1C1E", end_color="1C1C1E", fill_type="solid")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for d in docs:
        wo_id = d.get("work_order_id")
        proj_id = d.get("project_id")
        locked = wo_paid.get(wo_id, False) if wo_id else project_paid.get(proj_id, False) if proj_id else False
        ws.append([
            d.get("date", ""), d.get("vendor", ""), d.get("category", ""),
            projects.get(proj_id, ""), work_orders.get(wo_id, ""),
            int(round(float(d.get("amount") or 0))),
            "Ya" if d.get("is_billable") else "Tidak",
            "Ya" if locked else "Tidak",
            d.get("notes", ""),
        ])

    # Lebar kolom otomatis menyesuaikan konten, biar langsung enak dibaca tanpa perlu resize manual
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    # Format kolom Jumlah sebagai angka dengan pemisah ribuan, memudahkan bikin pivot table
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        row[0].number_format = "#,##0"

    ws.freeze_panes = "A2"  # header tetap kelihatan saat scroll panjang

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"pengeluaran_{datetime.now(timezone.utc).date().isoformat()}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@api_router.get("/reports/export")
async def export_csv():
    docs = await db.expenses.find().sort("date", -1).to_list(5000)
    projects = {p["id"]: p["name"] for p in await db.projects.find().to_list(1000)}
    work_orders = {w["id"]: w["name"] for w in await db.work_orders.find().to_list(2000)}

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Tanggal", "Vendor", "Kategori", "Proyek", "Work Order", "Jumlah (Rp)", "Billable", "Catatan"])
    for d in docs:
        writer.writerow([
            d.get("date", ""), d.get("vendor", ""), d.get("category", ""),
            projects.get(d.get("project_id"), ""), work_orders.get(d.get("work_order_id"), ""),
            int(round(float(d.get("amount") or 0))),
            "Ya" if d.get("is_billable") else "Tidak", d.get("notes", ""),
        ])
    output.seek(0)
    filename = f"pengeluaran_{datetime.now(timezone.utc).date().isoformat()}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _rl_image(raw: bytes, max_w_mm=80, max_h_mm=95):
    from reportlab.platypus import Image as RLImage
    from PIL import Image as PILImage
    im = PILImage.open(io.BytesIO(raw))
    if im.mode not in ("RGB", "L"):
        im = im.convert("RGB")
    im.thumbnail((1400, 1400))
    bio = io.BytesIO()
    im.save(bio, format="JPEG", quality=72)
    bio.seek(0)
    w, h = im.size
    ratio = min((max_w_mm * mm) / w, (max_h_mm * mm) / h)
    return RLImage(bio, width=w * ratio, height=h * ratio)


def _no_receipt_box(w_mm, h_mm, muted, border):
    t = Table([["Tanpa Struk"]], colWidths=[w_mm * mm], rowHeights=[h_mm * mm])
    t.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.6, border), ("TEXTCOLOR", (0, 0), (-1, -1), muted),
        ("FONTSIZE", (0, 0), (-1, -1), 9), ("BACKGROUND", (0, 0), (-1, -1), rl_colors.HexColor("#F5F5F3")),
    ]))
    return t


async def _gather_receipts(expenses):
    out = {}
    for e in expenses:
        p = e.get("receipt_path")
        if not p:
            continue
        try:
            content, _ = await run_in_threadpool(get_object, p)
            out[e["id"]] = content
        except Exception as ex:
            logger.error(f"receipt fetch failed {p}: {ex}")
    return out


def _build_expense_pdf(buf, title, subtitle_lines, groups, receipts, total_label="TOTAL", include_attachments=True):
    from reportlab.platypus import Image as RLImage
    styles = getSampleStyleSheet()
    ink = rl_colors.HexColor("#1C1C1E")
    moss = rl_colors.HexColor("#4A7C59")
    muted = rl_colors.HexColor("#8A8A8D")
    line = rl_colors.HexColor("#E5E5E3")

    h1 = ParagraphStyle("h1", parent=styles["Title"], textColor=ink, fontSize=22, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], textColor=muted, fontSize=10)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=ink, fontSize=13, spaceBefore=10, spaceAfter=4)
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=9, textColor=ink)
    cap = ParagraphStyle("cap", parent=styles["Normal"], fontSize=9, textColor=ink, spaceAfter=4)

    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=18 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm)
    elements = [Paragraph(title, h1)]
    for s in subtitle_lines:
        elements.append(Paragraph(s, sub))
    elements.append(Spacer(1, 8 * mm))

    def exp_table(exps):
        data = [["Tanggal", "Vendor", "Kategori", "Jumlah"]]
        subtotal = 0.0
        for e in exps:
            subtotal += float(e.get("amount") or 0)
            data.append([e.get("date", ""), Paragraph(e.get("vendor", ""), cell),
                         e.get("category", ""), rupiah(e.get("amount"))])
        t = Table(data, colWidths=[24 * mm, 70 * mm, 40 * mm, 34 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ink), ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"), ("LINEBELOW", (0, 1), (-1, -1), 0.4, line),
            ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        return t, subtotal

    grand = 0.0
    attach_exps = []
    for g in groups:
        elements.append(Paragraph(g["name"], h2))
        if g["expenses"]:
            t, st = exp_table(g["expenses"])
            elements.append(t)
            elements.append(Paragraph(f"<para align='right'><b>Subtotal: {rupiah(st)}</b></para>", cell))
            grand += st
            attach_exps.extend(g["expenses"])
        else:
            elements.append(Paragraph("Belum ada pengeluaran.", sub))
        elements.append(Spacer(1, 4 * mm))

    elements.append(Spacer(1, 6 * mm))
    total_tbl = Table([[total_label, rupiah(grand)]], colWidths=[134 * mm, 34 * mm])
    total_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), moss), ("TEXTCOLOR", (0, 0), (-1, -1), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"), ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8), ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(total_tbl)

    attach = attach_exps if include_attachments else []
    if attach:
        elements.append(Spacer(1, 8 * mm))
        with_receipt_count = sum(1 for e in attach if receipts.get(e["id"]))
        elements.append(Paragraph(f"Lampiran Bukti Struk ({with_receipt_count} dari {len(attach)} transaksi)", h2))

        # Grid 2 kolom supaya beberapa struk muat dalam satu halaman (target 4-6/halaman)
        # tapi ukuran gambar tetap cukup besar untuk tetap terbaca.
        COLS = 2
        page_content_w = 178  # mm: A4 (210) dikurangi leftMargin+rightMargin (16+16)
        cell_w_mm = page_content_w / COLS
        img_max_w = cell_w_mm - 10  # sisakan ruang untuk padding kiri-kanan
        img_max_h = 62  # dikecilkan supaya 3 baris (6 struk) muat per halaman penuh

        groups_with_exp = [g for g in groups if g["expenses"]]
        show_divider = len(groups_with_exp) > 1  # cuma perlu pembatas kalau lampiran gabungan >1 WO

        for gi, g in enumerate(groups_with_exp):
            if show_divider:
                if gi > 0:
                    elements.append(Spacer(1, 3 * mm))
                wo_bar = Table([[g["name"]]], colWidths=[page_content_w * mm])
                wo_bar.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), rl_colors.HexColor("#EFEDE6")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), ink),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8), ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]))
                elements.append(wo_bar)
                elements.append(Spacer(1, 3 * mm))

            cells = []
            for e in g["expenses"]:
                caption = Paragraph(
                    f"{e.get('vendor','')} — {e.get('date','')}<br/><b>{rupiah(e.get('amount'))}</b>", cap)
                raw = receipts.get(e["id"])
                visual = None
                if raw:
                    try:
                        visual = _rl_image(raw, max_w_mm=img_max_w, max_h_mm=img_max_h)
                    except Exception as ex:
                        logger.error(f"embed receipt failed: {ex}")
                if visual is None:
                    visual = _no_receipt_box(img_max_w, img_max_h, muted, line)
                cells.append([visual, Spacer(1, 2 * mm), caption])

            while len(cells) % COLS != 0:
                cells.append([Spacer(1, 1)])

            rows = [cells[i:i + COLS] for i in range(0, len(cells), COLS)]
            grid = Table(rows, colWidths=[cell_w_mm * mm] * COLS)
            grid.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("LINEBELOW", (0, 0), (-1, -2), 0.4, line),
            ]))
            elements.append(grid)

    doc.build(elements)


@api_router.get("/projects/{project_id}/pdf")
async def project_pdf(project_id: str, work_order_ids: Optional[str] = Query(None)):
    proj = await db.projects.find_one({"id": project_id})
    if not proj:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")

    work_orders = await db.work_orders.find({"project_id": project_id}).sort("created_at", 1).to_list(2000)
    all_exp = await db.expenses.find({"project_id": project_id}).sort("date", 1).to_list(5000)

    # work_order_ids: daftar id WO yang dipilih user (pisah koma), plus sentinel
    # "__unassigned__" untuk menyertakan pengeluaran tanpa WO. Kalau parameter ini
    # tidak dikirim sama sekali, default-nya sertakan semua (perilaku lama).
    include_unassigned = True
    if work_order_ids is not None:
        selected = set(x for x in work_order_ids.split(",") if x)
        include_unassigned = "__unassigned__" in selected
        work_orders = [w for w in work_orders if w["id"] in selected]

    groups = []
    for w in work_orders:
        groups.append({"name": f"Work Order: {w['name']}",
                       "expenses": [e for e in all_exp if e.get("work_order_id") == w["id"]]})
    unassigned = [e for e in all_exp if not e.get("work_order_id")]
    if unassigned and include_unassigned:
        groups.append({"name": "Tanpa Work Order", "expenses": unassigned})

    included_exp = [e for g in groups for e in g["expenses"]]
    receipts = await _gather_receipts(included_exp)
    subtitle = [f"Proyek: <b>{proj['name']}</b>"]
    if proj.get("client"):
        subtitle.append(f"Klien: {proj['client']}")
    subtitle.append(f"Dibuat: {datetime.now(timezone.utc).strftime('%d %b %Y')}")

    buf = io.BytesIO()
    _build_expense_pdf(buf, "LAPORAN PENGELUARAN", subtitle, groups, receipts, "TOTAL PROYEK")
    buf.seek(0)
    safe = "".join(ch for ch in proj["name"] if ch.isalnum() or ch in (" ", "-", "_")).strip().replace(" ", "_")
    return StreamingResponse(iter([buf.getvalue()]), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=laporan_{safe or 'proyek'}.pdf"})


@api_router.get("/projects/{project_id}/invoice")
async def project_invoice(project_id: str, work_order_ids: Optional[str] = Query(None)):
    """Dokumen tagihan internal ke tim Finance — cuma berisi pengeluaran yang
    ditandai 'Bisa ditagihkan' (is_billable), tanpa lampiran foto struk (beda
    dari laporan pengeluaran biasa yang menyertakan semua transaksi + bukti)."""
    proj = await db.projects.find_one({"id": project_id})
    if not proj:
        raise HTTPException(status_code=404, detail="Proyek tidak ditemukan")

    work_orders = await db.work_orders.find({"project_id": project_id}).sort("created_at", 1).to_list(2000)
    all_exp = await db.expenses.find({"project_id": project_id, "is_billable": True}).sort("date", 1).to_list(5000)

    include_unassigned = True
    if work_order_ids is not None:
        selected = set(x for x in work_order_ids.split(",") if x)
        include_unassigned = "__unassigned__" in selected
        work_orders = [w for w in work_orders if w["id"] in selected]

    groups = []
    for w in work_orders:
        groups.append({"name": f"Work Order: {w['name']}",
                       "expenses": [e for e in all_exp if e.get("work_order_id") == w["id"]]})
    unassigned = [e for e in all_exp if not e.get("work_order_id")]
    if unassigned and include_unassigned:
        groups.append({"name": "Tanpa Work Order", "expenses": unassigned})

    total_billable = sum(e.get("amount", 0) for g in groups for e in g["expenses"])
    if total_billable == 0:
        raise HTTPException(status_code=400, detail="Tidak ada pengeluaran 'Bisa ditagihkan' pada proyek/WO yang dipilih.")

    subtitle = [f"Proyek: <b>{proj['name']}</b>"]
    if proj.get("client"):
        subtitle.append(f"Klien: {proj['client']}")
    subtitle.append(f"Dibuat: {datetime.now(timezone.utc).strftime('%d %b %Y')}")
    subtitle.append("Hanya memuat pengeluaran yang ditandai <b>Bisa Ditagihkan</b>")

    buf = io.BytesIO()
    _build_expense_pdf(buf, "TAGIHAN KE FINANCE", subtitle, groups, {}, "TOTAL TAGIHAN", include_attachments=False)
    buf.seek(0)
    safe = "".join(ch for ch in proj["name"] if ch.isalnum() or ch in (" ", "-", "_")).strip().replace(" ", "_")
    return StreamingResponse(iter([buf.getvalue()]), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=tagihan_{safe or 'proyek'}.pdf"})


@api_router.get("/work-orders/{wo_id}/pdf")
async def work_order_pdf(wo_id: str):
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order tidak ditemukan")
    proj = await db.projects.find_one({"id": wo.get("project_id")})
    exps = await db.expenses.find({"work_order_id": wo_id}).sort("date", 1).to_list(5000)

    groups = [{"name": f"Work Order: {wo['name']}", "expenses": exps}]
    receipts = await _gather_receipts(exps)
    subtitle = []
    if proj:
        subtitle.append(f"Proyek: <b>{proj['name']}</b>")
    subtitle.append(f"Work Order: {wo['name']}")
    subtitle.append(f"Dibuat: {datetime.now(timezone.utc).strftime('%d %b %Y')}")

    buf = io.BytesIO()
    _build_expense_pdf(buf, "LAPORAN WORK ORDER", subtitle, groups, receipts, "TOTAL PENGELUARAN")
    buf.seek(0)
    safe = "".join(ch for ch in wo["name"] if ch.isalnum() or ch in (" ", "-", "_")).strip().replace(" ", "_")
    return StreamingResponse(iter([buf.getvalue()]), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=laporan_wo_{safe or 'wo'}.pdf"})


BACKUP_COLLECTIONS = ["categories", "projects", "work_orders", "expenses"]


def _strip_mongo_id(doc: dict) -> dict:
    doc = dict(doc)
    doc.pop("_id", None)
    return doc


@api_router.get("/backup")
async def create_backup():
    """Export semua data (kategori, proyek, work order, pengeluaran) + file struk
    jadi satu file .zip yang bisa disimpan di luar VPS (Google Drive, HP, dst)."""
    data = {"version": 1, "exported_at": now_iso(), "app": APP_NAME}
    for coll in BACKUP_COLLECTIONS:
        docs = await db[coll].find().to_list(100000)
        data[coll] = [_strip_mongo_id(d) for d in docs]

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("data.json", json.dumps(data, ensure_ascii=False, indent=2))
        # Sertakan semua file struk yang ada di disk. Path di dalam zip disamakan
        # dengan yang tersimpan di field receipt_path supaya gampang dipetakan balik.
        if UPLOAD_DIR.exists():
            for f in UPLOAD_DIR.rglob("*"):
                if f.is_file():
                    rel = f.relative_to(UPLOAD_DIR)
                    zf.write(f, arcname=f"files/{rel.as_posix()}")

    buf.seek(0)
    filename = f"easy-expense-backup-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.zip"
    return StreamingResponse(iter([buf.getvalue()]), media_type="application/zip",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@api_router.post("/restore")
async def restore_backup(file: UploadFile = File(...)):
    """Import balik file backup .zip.

    Aman dijalankan berkali-kali / restore backup yang sama dua kali: setiap
    dokumen di-upsert berdasarkan field 'id' (bukan insert baru), jadi tidak
    akan menggandakan data. Data yang ADA SEKARANG tapi TIDAK ADA di file
    backup tetap dibiarkan apa adanya (restore ini menggabungkan/memperbarui,
    bukan menghapus/menimpa total)."""
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="File backup kosong")

    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="File bukan .zip backup yang valid")

    try:
        data = json.loads(zf.read("data.json"))
    except KeyError:
        raise HTTPException(status_code=400, detail="File backup tidak lengkap (data.json tidak ditemukan)")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="File backup rusak (data.json tidak bisa dibaca)")

    summary = {}
    for coll in BACKUP_COLLECTIONS:
        docs = data.get(coll, [])
        count = 0
        for doc in docs:
            doc_id = doc.get("id")
            if not doc_id:
                continue
            doc = _strip_mongo_id(doc)
            await db[coll].update_one({"id": doc_id}, {"$set": doc}, upsert=True)
            count += 1
        summary[coll] = count

    files_restored = 0
    for name in zf.namelist():
        if name.startswith("files/") and not name.endswith("/"):
            rel = name[len("files/"):]
            dest = (UPLOAD_DIR / rel).resolve()
            if not str(dest).startswith(str(UPLOAD_DIR.resolve())):
                continue  # jaga-jaga kalau nama file di zip aneh/path traversal
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(zf.read(name))
            files_restored += 1

    summary["files"] = files_restored
    return {"ok": True, "restored": summary}


@api_router.get("/")
async def root():
    return {"message": "Easy Expense API", "status": "ok"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


async def ensure_category_order():
    # Migrasi: kategori lama (dibuat sebelum fitur urutan-manual ini ada) belum punya
    # field 'order' -> isi berdasarkan created_at supaya urutan yang sudah ada tidak berubah.
    missing = await db.categories.count_documents({"order": {"$exists": False}})
    if missing > 0:
        docs = await db.categories.find().sort("created_at", 1).to_list(1000)
        for i, d in enumerate(docs):
            await db.categories.update_one({"id": d["id"]}, {"$set": {"order": i}})


@app.on_event("startup")
async def startup():
    try:
        await ensure_default_categories()
        await ensure_category_order()
        logger.info("Default categories ensured")
    except Exception as e:
        logger.error(f"Category seed failed: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
